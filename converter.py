"""
converter.py — Professional File Parsing and Enterprise Excel Export Logic.
Deep Learning Edition: Transformer Embeddings -> Siamese FFNN -> Graph CC
"""

from __future__ import annotations

import io
import json
import logging
import os
import unicodedata
import re
import warnings
from abc import ABC, abstractmethod
from copy import copy
from pathlib import Path
from typing import Dict, Type, Tuple, List, Optional

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string

# Deep Learning & Graph Imports
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import TensorDataset, DataLoader
from sentence_transformers import SentenceTransformer
import networkx as nx
from sklearn.metrics import average_precision_score, f1_score, precision_recall_curve, precision_score, recall_score, roc_curve

# Setup professional logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  Custom Exceptions & Pytorch Models
# ─────────────────────────────────────────────────────────────────────────────

class ConversionError(Exception):
    pass

class UnsupportedFileTypeError(ConversionError):
    pass

class EmptyFileError(ConversionError):
    pass

class SiameseZoneClassifier(nn.Module):
    """Feed-Forward Neural Network with Categorical Zone Embeddings."""
    def __init__(self, num_zones: int = 1, zone_dim: int = 32, text_dim: int = 384):
        super().__init__()
        self.zone_embedding = nn.Embedding(num_zones, zone_dim)
        row_dim = text_dim + zone_dim
        
        self.ffnn = nn.Sequential(
            nn.LayerNorm(row_dim * 3),
            nn.Linear(row_dim * 3, 512),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(512, 256),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(256, 1)
        )

    def forward(self, u_text: torch.Tensor, u_zone: torch.Tensor, v_text: torch.Tensor, v_zone: torch.Tensor) -> torch.Tensor:
        u_z = self.zone_embedding(u_zone)
        v_z = self.zone_embedding(v_zone)
        
        u = torch.cat([u_text, u_z], dim=1)
        v = torch.cat([v_text, v_z], dim=1)
        
        diff = torch.abs(u - v)
        features = torch.cat([u, v, diff], dim=1)
        return self.ffnn(features)

# ─────────────────────────────────────────────────────────────────────────────
#  Parser Strategy Pattern (Intact)
# ─────────────────────────────────────────────────────────────────────────────

class BaseParser(ABC):
    @abstractmethod
    def parse(self, path: Path) -> Tuple[pd.DataFrame, str]:
        pass

class PDFParser(BaseParser):
    def parse(self, path: Path) -> Tuple[pd.DataFrame, str]:
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber is required for PDF parsing. Run: pip install pdfplumber")

        tables, text_rows, pages_read = [], [], 0
        with pdfplumber.open(path) as pdf:
            if not pdf.pages:
                raise ConversionError("The PDF has no pages.")
            for page in pdf.pages:
                pages_read += 1
                page_tables = page.extract_tables()
                if page_tables:
                    for table in page_tables:
                        if table and len(table) > 1:
                            tables.append(pd.DataFrame(table[1:], columns=table[0]))
                else:
                    text = page.extract_text()
                    if text:
                        text_rows.extend(text.splitlines())

        if tables:
            return pd.concat(tables, ignore_index=True), f"Extracted tables from {pages_read} PDF page(s)"
        if text_rows:
            df = pd.DataFrame({"Content": [line.strip() for line in text_rows if line.strip()]})
            return df, f"Extracted raw text from {pages_read} PDF page(s)"
        
        raise EmptyFileError("No extractable text or tables found in PDF (might be a scanned image).")

class ExcelParser(BaseParser):
    def parse(self, path: Path) -> Tuple[pd.DataFrame, str]:
        try:
            xl = pd.ExcelFile(path)
        except Exception as e:
            raise ConversionError(f"Failed to open Excel: {e}")

        sheet_names = xl.sheet_names
        if len(sheet_names) == 1:
            raw = xl.parse(sheet_names[0], header=None)
            df, info = self._flatten_repeating_headers(raw)
            if df is None:
                df = xl.parse(sheet_names[0])
                info = f"Loaded sheet '{sheet_names[0]}'"
            return df, info
        else:
            frames = []
            for name in sheet_names:
                raw = xl.parse(name, header=None)
                frame, _ = self._flatten_repeating_headers(raw)
                if frame is None:
                    frame = xl.parse(name)
                frame.insert(0, "Sheet_Source", name)
                frames.append(frame)
            return pd.concat(frames, ignore_index=True), f"Merged {len(sheet_names)} sheets"

    def _flatten_repeating_headers(self, raw: pd.DataFrame) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
        header_mask = raw[0].astype(str).str.strip().eq("Date")
        header_indices = raw.index[header_mask].tolist()

        if len(header_indices) < 2:
            return None, None

        canonical_cols = [str(v).strip() for v in raw.loc[header_indices[0]].tolist()]
        blocks = []

        for i, h_idx in enumerate(header_indices):
            next_h = header_indices[i + 1] if i + 1 < len(header_indices) else len(raw)
            block = raw.iloc[h_idx + 1 : next_h].copy()
            if not block.empty:
                block.columns = canonical_cols
                blocks.append(block)

        if not blocks:
            return None, None

        df = pd.concat(blocks, ignore_index=True)
        return df, f"Flattened {len(header_indices)} repeating header blocks"

class TextCSVParser(BaseParser):
    def parse(self, path: Path) -> Tuple[pd.DataFrame, str]:
        content = None
        for enc in ["utf-8", "utf-16", "latin-1", "cp1252"]:
            try:
                content = path.read_text(encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise ConversionError("Unable to decode text file with standard encodings.")

        for sep in [",", "\t", ";", "|"]:
            try:
                df = pd.read_csv(io.StringIO(content), sep=sep, engine="python")
                if len(df.columns) > 1:
                    return df, f"Parsed as CSV (separator='{sep}')"
            except:
                continue
        
        lines = content.splitlines()
        df = pd.DataFrame({"Content": [l.strip() for l in lines if l.strip()]})
        return df, "Parsed as raw text"

# ─────────────────────────────────────────────────────────────────────────────
#  Main Engine
# ─────────────────────────────────────────────────────────────────────────────

class MergeGroupReader:
    def __init__(self, path: str | Path, column_letter: str = "H"):
        self.path = Path(path)
        self.column_letter = column_letter.upper()

    def read(self) -> Dict[Tuple[str, int], Optional[int]]:
        if not self.path.exists():
            raise FileNotFoundError(f"File not found: {self.path}")

        wb = openpyxl.load_workbook(self.path, read_only=False, data_only=True)
        target_col = column_index_from_string(self.column_letter)

        row_to_group: Dict[Tuple[str, int], Optional[int]] = {}
        global_group_id = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx in range(1, ws.max_row + 1):
                row_to_group[(sheet_name, row_idx)] = None

            merged_ranges = [
                merge_range for merge_range in ws.merged_cells.ranges
                if merge_range.min_col == target_col and merge_range.max_col == target_col
            ]
            merged_ranges.sort(key=lambda rng: (rng.min_row, rng.max_row))

            for merge_range in merged_ranges:
                for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                    row_to_group[(sheet_name, row_idx)] = global_group_id
                global_group_id += 1

        return row_to_group

class FileConverter:
    _PARSER_MAP: Dict[str, Type[BaseParser]] = {
        ".pdf": PDFParser,
        ".xlsx": ExcelParser,
        ".xls": ExcelParser,
        ".csv": TextCSVParser,
        ".txt": TextCSVParser,
    }

    def __init__(self, trip_type: str = "ramassage", max_passengers: int = 4, similarity_threshold: float = 0.5):
        self.trip_label = "Ramassage" if trip_type.lower() == "ramassage" else "Retour"
        self.max_passengers = int(max_passengers)
        self.pairwise_threshold = float(similarity_threshold)
        self.similarity_threshold = self.pairwise_threshold

        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        logger.info(f"Loading DL Encoders on {self.device}...")
        
        # Categorical state tracking
        self.zone_to_idx: Dict[str, int] = {"<UNKNOWN>": 0}
        self.num_zones = 1
        
        paths = self._artifact_paths()
        if paths.get("encoder_dir") and paths["encoder_dir"].exists():
            logger.info("Loading fine-tuned SentenceTransformer...")
            self.encoder = SentenceTransformer(str(paths["encoder_dir"]), device=self.device)
        else:
            self.encoder = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2", device=self.device)
            
        self.classifier = SiameseZoneClassifier(num_zones=self.num_zones).to(self.device)

        self.reference_col_widths = [14.855, 20.14, 26.0, 23.42, 17.57, 62.85, 19.28, 16.0]
        self.reference_row_height = 15.75
        self.reference_header_fill = "FF000000"
        self.reference_header_text = "FFFF7E00"
        self.reference_taxi_header_text = "FFFFFF00"
        self.reference_course_orange = "FFFF9900"
        self.reference_course_white = "FFFFFFFF"
        self.reference_font_size = 11

        self._load_artifacts()

    def convert(self, input_path: str, output_path: str) -> str:
        path = Path(input_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {input_path}")
        ext = path.suffix.lower()
        if ext not in self._PARSER_MAP:
            raise UnsupportedFileTypeError(f"Extension {ext} is not supported.")
        parser = self._PARSER_MAP[ext]()
        df, info = parser.parse(path)
        df = self._normalize_dataframe(df)
        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"
        self._export_to_styled_excel(df, output_path)
        return f"Success: {info}. Saved to {output_path}"

    def convert_dispatch_ml(self, input_path: str, output_path: str) -> str:
        path = Path(input_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {input_path}")

        raw_df = self._load_structured_input(path)
        cleaned_df = self._remove_repeated_headers(raw_df)
        ml_df = self._prepare_ml_dataframe(cleaned_df)
        
        row_vectors = self._vectorize_addresses(ml_df)
        clusters = self._build_pairwise_clusters(ml_df, row_vectors)
        self._build_dispatch_outputs(ml_df, clusters)

        export_cleaned_df = cleaned_df.copy()
        for col in ["_excel_row", "_sheet_name"]:
            if col in export_cleaned_df.columns:
                export_cleaned_df = export_cleaned_df.drop(columns=[col])

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        self._export_dispatch_workbook(path, export_cleaned_df, clusters, output_path)
        return f"Success: Dispatch workbook saved to {output_path}"

    def _normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.dropna(how="all")
        df = df[~df.apply(lambda r: r.astype(str).str.strip().eq("").all(), axis=1)]
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].astype(str).str.strip()
        if len(df.columns) >= 3:
            cols = list(df.columns)
            cols[2] = self.trip_label
            df.columns = cols
        return df.reset_index(drop=True)

    def _load_structured_input(self, path: Path) -> pd.DataFrame:
        if path.suffix.lower() in {".xlsx", ".xls"}:
            xl = pd.ExcelFile(path)
            frames: List[pd.DataFrame] = []
            for sheet_name in xl.sheet_names:
                frame = xl.parse(sheet_name)
                frame["_excel_row"] = list(range(2, len(frame) + 2))
                frame["_sheet_name"] = sheet_name
                if len(xl.sheet_names) > 1:
                    frame.insert(0, "Sheet_Source", sheet_name)
                frames.append(frame)
            if not frames:
                raise EmptyFileError("Excel workbook has no readable sheets.")
            return pd.concat(frames, ignore_index=True)
        frame = pd.read_csv(path)
        frame["_excel_row"] = list(range(2, len(frame) + 2))
        frame["_sheet_name"] = "__csv__"
        return frame

    def _normalize_colname(self, name: str) -> str:
        raw = str(name).strip().lower()
        normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
        return "".join(ch for ch in normalized if ch.isalnum())

    def _resolve_required_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        normalized = {self._normalize_colname(col): col for col in df.columns}
        candidates = {
            "Date": ["date"],
            "Heure": ["heure", "time", "heuredarrivee", "arrivaltime"],
            "Ramassage": ["ramassage", "pickup", "pickuplocation", "adresse", "address"],
            "Destination": ["destination", "dropoff", "dropofflocation", "site"],
            "Passenger": ["nomprenom", "passenger", "passagers", "name"],
            "Zone": ["zone", "secteur", "region"],
        }
        resolved = {}
        fallback_order = list(df.columns)
        for i, (field, aliases) in enumerate(candidates.items()):
            match = next((normalized[key] for key in aliases if key in normalized), None)
            if match is None:
                match = fallback_order[min(i, len(fallback_order) - 1)] if field != "Zone" else list(df.columns)[0]
            resolved[field] = match
        return resolved

    def _resolve_optional_column(self, df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
        normalized = {self._normalize_colname(col): col for col in df.columns}
        for alias in aliases:
            key = self._normalize_colname(alias)
            if key in normalized:
                return normalized[key]
        return None

    def _remove_repeated_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            raise EmptyFileError("Input file is empty.")
        header_rows = []
        for idx, row in df.iterrows():
            if all(str(row[col]).strip() == str(col).strip() for col in df.columns):
                header_rows.append(idx)
        bad_rows = sorted(set(header_rows))
        cleaned_df = df.drop(index=bad_rows)
        if cleaned_df.empty:
            raise EmptyFileError("No usable rows remain after removing repeated headers.")
        return cleaned_df

    def _prepare_ml_dataframe(self, cleaned_df: pd.DataFrame, row_to_group: Optional[Dict] = None) -> pd.DataFrame:
        working_df = cleaned_df.copy()
        matricule_col = self._resolve_optional_column(working_df, ["matricule"])
        if matricule_col is not None:
            working_df = working_df.drop(columns=[matricule_col])

        cols = self._resolve_required_columns(working_df)
        operation_col = self._resolve_optional_column(working_df, ["operation"])
        sheet_source_col = self._resolve_optional_column(working_df, ["sheet_source"])

        zone_col_name = cols.get("Zone")
        zone_data = working_df[zone_col_name] if zone_col_name in working_df.columns else pd.Series("UNKNOWN", index=working_df.index)

        ml_df = pd.DataFrame({
            "Date": working_df[cols["Date"]],
            "Heure": working_df[cols["Heure"]],
            "Ramassage": working_df[cols["Ramassage"]],
            "Destination": working_df[cols["Destination"]],
            "Nom - Prénom": working_df[cols["Passenger"]],
            "Zone": zone_data,
        }).copy()

        ml_df["Opération"] = working_df[operation_col] if operation_col is not None else ""
        if "_excel_row" in working_df.columns:
            ml_df["_excel_row"] = working_df["_excel_row"].values
        if "_sheet_name" in working_df.columns:
            ml_df["_sheet_name"] = working_df["_sheet_name"].astype(str).values
        if sheet_source_col is not None:
            ml_df["Sheet_Source"] = working_df[sheet_source_col].astype(str)

        for col in ["Ramassage", "Destination", "Nom - Prénom", "Opération"]:
            ml_df[col] = ml_df[col].fillna("").astype(str).str.strip()

        ml_df["Zone"] = ml_df["Zone"].fillna("UNKNOWN").astype(str).str.strip().str.upper()

        ml_df["Date"] = pd.to_datetime(ml_df["Date"], errors="coerce", dayfirst=True)
        ml_df["Date"] = ml_df["Date"].fillna(pd.Timestamp("1970-01-01"))
        ml_df["dispatch_date"] = ml_df["Date"].dt.date

        time_features = self._extract_time_features(ml_df["Heure"])
        ml_df["dispatch_time_key"] = time_features["time_key"]
        ml_df["hour"] = time_features["hour"].fillna(0).astype(int)
        ml_df["Heure"] = pd.to_datetime(ml_df["dispatch_time_key"], format="%H:%M", errors="coerce").dt.time
        ml_df["weekday"] = ml_df["Date"].dt.weekday

        ml_df["Ramassage_clean"] = ml_df["Ramassage"].str.lower()
        ml_df["Destination_clean"] = ml_df["Destination"].str.lower()
        ml_df["Ramassage_normalized"] = ml_df["Ramassage"].apply(self._normalize_address)
        ml_df["Destination_normalized"] = ml_df["Destination"].apply(self._normalize_address)
        ml_df["route"] = ml_df["Ramassage_clean"] + " > " + ml_df["Destination_clean"]
        ml_df["operation_clean"] = ml_df["Opération"].str.lower()

        if row_to_group is not None:
            group_ids = [
                row_to_group.get((str(row.get("_sheet_name", "__csv__")), int(row.get("_excel_row", -1))))
                for _, row in ml_df.iterrows()
            ]
            ml_df["merge_group_id"] = group_ids
            ml_df["pickup_zone_candidate"] = group_ids
            ml_df["destination_zone_candidate"] = group_ids
        else:
            ml_df["merge_group_id"] = pd.NA
            ml_df["pickup_zone_candidate"] = pd.NA
            ml_df["destination_zone_candidate"] = pd.NA

        return ml_df

    def _extract_time_features(self, heure_series: pd.Series) -> pd.DataFrame:
        def parse_time_value(value) -> Tuple[float, str]:
            if pd.isna(value): return np.nan, "00:00"
            if hasattr(value, "hour"):
                hour, minute = int(value.hour), int(getattr(value, "minute", 0) or 0)
                return float(hour), f"{hour:02d}:{minute:02d}"
            text = str(value).strip().lower()
            match_hm = pd.Series(text).str.extract(r"(\d{1,2})\s*[:h]\s*(\d{1,2})", expand=True).iloc[0]
            if match_hm.notna().all():
                return float(match_hm[0]), f"{int(match_hm[0]):02d}:{int(match_hm[1]):02d}"
            return np.nan, "00:00"

        parsed = heure_series.apply(parse_time_value)
        return pd.DataFrame({"hour": parsed.apply(lambda t: t[0]), "time_key": parsed.apply(lambda t: t[1])})

    def _normalize_address(self, value: str) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r"[^0-9a-z\s-]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def _artifact_dir(self) -> Path:
        return Path(__file__).resolve().parent / "data"

    def _artifact_paths(self) -> Dict[str, Path]:
        base = self._artifact_dir()
        return {
            "siamese_head": base / "siamese_head.pth",
            "threshold": base / "pairwise_threshold.json",
            "encoder_dir": base / "finetuned_encoder",
            "zone_map": base / "zone_to_idx.json",
        }

    def _load_artifacts(self) -> None:
        paths = self._artifact_paths()
        try:
            if paths["zone_map"].exists():
                self.zone_to_idx = json.loads(paths["zone_map"].read_text(encoding="utf-8"))
                self.num_zones = len(self.zone_to_idx)
                self.classifier = SiameseZoneClassifier(num_zones=self.num_zones).to(self.device)
                
            if paths["siamese_head"].exists():
                self.classifier.load_state_dict(torch.load(paths["siamese_head"], map_location=self.device))
                self.classifier.eval()
                
            if paths["threshold"].exists():
                payload = json.loads(paths["threshold"].read_text(encoding="utf-8"))
                self.pairwise_threshold = self._sanitize_pairwise_threshold(payload.get("pairwise_threshold", 0.5))
                self.similarity_threshold = self.pairwise_threshold
        except Exception as exc:
            logger.warning("Failed to load persisted artifacts: %s", exc)

    def _save_artifacts(self, output_dir: str = "data") -> None:
        artifact_dir = Path(output_dir)
        artifact_dir.mkdir(parents=True, exist_ok=True)
        paths = self._artifact_paths()
        
        torch.save(self.classifier.state_dict(), paths["siamese_head"])
        paths["threshold"].write_text(
            json.dumps({"pairwise_threshold": float(self.pairwise_threshold)}, ensure_ascii=False),
            encoding="utf-8"
        )
        paths["zone_map"].write_text(
            json.dumps(self.zone_to_idx, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        self.encoder.save(str(paths["encoder_dir"]))

    def _sanitize_pairwise_threshold(self, value: float) -> float:
        try:
            threshold = float(value)
            return float(np.clip(threshold, 0.0, 1.0)) if np.isfinite(threshold) else 0.5
        except (TypeError, ValueError):
            return 0.5

    def _vectorize_addresses(self, ml_df: pd.DataFrame, fit: bool = False) -> torch.Tensor:
        """Deep Learning approach: Pass combined string properties through the transformer"""
        texts = self._build_row_texts(ml_df)
        if not texts:
            return torch.empty((0, 384), dtype=torch.float32).to(self.device)

        with torch.no_grad():
            embeddings = self.encoder.encode(texts, convert_to_tensor=True, show_progress_bar=False)
            
        return embeddings

    def _build_row_texts(self, ml_df: pd.DataFrame) -> List[str]:
        # Only the pickup address is a learning signal.
        # Destination, route, and operation are intentionally excluded.
        return ml_df["Ramassage_normalized"].fillna("").astype(str).tolist()

    def _build_pairwise_examples(self, ml_df: pd.DataFrame, row_vectors: torch.Tensor, row_to_group: Dict) -> pd.DataFrame:
        """Updated: 50% Hard Negatives, 50% Random Negatives to stabilize global geometry"""
        group_ids = [
            row_to_group.get((str(row.get("_sheet_name", "__csv__")), int(row.get("_excel_row", -1))))
            for _, row in ml_df.iterrows()
        ]
        
        groups, bucket_map = {}, {}
        for idx, (gid, (_, row)) in enumerate(zip(group_ids, ml_df.iterrows())):
            if gid is not None:
                groups.setdefault(int(gid), []).append(idx)
                bucket_key = (row["dispatch_date"], row["dispatch_time_key"])
                bucket_map.setdefault(bucket_key, []).append(idx)

        all_labeled = [i for i, g in enumerate(group_ids) if g is not None]
        records = []
        rng = np.random.default_rng(42)

        for gid, positions in groups.items():
            if len(positions) < 2: continue
            
            # Positives
            for left_idx in range(len(positions)):
                for right_idx in range(left_idx + 1, len(positions)):
                    a, b = positions[left_idx], positions[right_idx]
                    records.append({"left_idx": a, "right_idx": b, "label": 1})

                    # 1. Hard negative (same time bucket, different group)
                    anchor_bucket = (ml_df.iloc[a]["dispatch_date"], ml_df.iloc[a]["dispatch_time_key"])
                    same_bucket = [i for i in bucket_map.get(anchor_bucket, []) if group_ids[i] is not None and int(group_ids[i]) != int(gid)]
                    
                    if same_bucket:
                        anchor_vec = row_vectors[a].unsqueeze(0)
                        target_vecs = row_vectors[same_bucket]
                        cosine_scores = torch.nn.functional.cosine_similarity(anchor_vec, target_vecs).cpu().numpy()
                        
                        top_idx = int(rng.choice(np.argsort(cosine_scores)[::-1][:min(8, len(same_bucket))]))
                        records.append({"left_idx": a, "right_idx": same_bucket[top_idx], "label": 0})
                        
                    # 2. Easy/Random negative (any different group globally)
                    # This prevents the model from hallucinating false positives on easy pairs
                    other = [i for i in all_labeled if group_ids[i] is not None and int(group_ids[i]) != int(gid)]
                    if other:
                        records.append({"left_idx": a, "right_idx": int(rng.choice(other)), "label": 0})

        return pd.DataFrame.from_records(records)

    def _fit_pairwise_model(self, ml_df: pd.DataFrame, row_vectors: torch.Tensor, row_to_group: Dict) -> Dict:
        # --- VOCABULARY GENERATION ---
        unique_zones = sorted(list(ml_df["Zone"].unique()))
        self.zone_to_idx = {"<UNKNOWN>": 0}
        for idx, zone in enumerate(unique_zones, start=1):
            self.zone_to_idx[zone] = idx
        self.num_zones = len(self.zone_to_idx)
        
        # Re-initialize the classifier with the new vocabulary size
        self.classifier = SiameseZoneClassifier(num_zones=self.num_zones).to(self.device)
        
        zone_ints = [self.zone_to_idx.get(z, 0) for z in ml_df["Zone"]]
        zone_tensors = torch.tensor(zone_ints, dtype=torch.long)
        # -----------------------------

        pair_examples = self._build_pairwise_examples(ml_df, row_vectors, row_to_group)
        if pair_examples.empty:
            raise ConversionError("No labeled pairwise examples.")

        u_texts = row_vectors[pair_examples["left_idx"].values]
        v_texts = row_vectors[pair_examples["right_idx"].values]
        u_zones = zone_tensors[pair_examples["left_idx"].values]
        v_zones = zone_tensors[pair_examples["right_idx"].values]
        labels = torch.tensor(pair_examples["label"].values, dtype=torch.float32).unsqueeze(-1)

        dataset = TensorDataset(u_texts, v_texts, u_zones, v_zones, labels)
        loader = DataLoader(dataset, batch_size=64, shuffle=True)

        optimizer = optim.Adam(self.classifier.parameters(), lr=1e-3)
        criterion = nn.BCEWithLogitsLoss()

        logger.info(f"Phase 1: Training Siamese FFNN & Zone Embeddings on frozen text for 15 epochs...")
        self.classifier.train()
        epochs_frozen = 15
        
        for epoch in range(epochs_frozen):
            total_loss = 0
            for u_t, v_t, u_z, v_z, y in loader:
                u_t, v_t = u_t.to(self.device), v_t.to(self.device)
                u_z, v_z = u_z.to(self.device), v_z.to(self.device)
                y = y.to(self.device)
                
                optimizer.zero_grad()
                logits = self.classifier(u_t, u_z, v_t, v_z)
                loss = criterion(logits, y)
                loss.backward()
                optimizer.step()
                total_loss += loss.item()

        logger.info("Phase 2: Joint fine-tuning of transformer (last 2 layers) and classifier...")
        
        for param in self.encoder.parameters():
            param.requires_grad = False
            
        for name, param in self.encoder.named_parameters():
            if "layer.10" in name or "layer.11" in name or "pooler" in name or "Modules.1" in name:
                param.requires_grad = True

        encoder_params = filter(lambda p: p.requires_grad, self.encoder.parameters())
        
        optimizer_joint = optim.Adam([
            {'params': self.classifier.parameters(), 'lr': 1e-4}, 
            {'params': encoder_params, 'lr': 2e-5}                
        ])

        self.encoder.train()
        self.classifier.train()

        texts = self._build_row_texts(ml_df)
        u_raw_texts = [texts[i] for i in pair_examples["left_idx"].values]
        v_raw_texts = [texts[i] for i in pair_examples["right_idx"].values]
        
        batch_size = 64
        num_samples = len(u_raw_texts)
        epochs_joint = 10

        for epoch in range(epochs_joint):
            indices = torch.randperm(num_samples).tolist()
            total_loss = 0
            for start_idx in range(0, num_samples, batch_size):
                batch_idx = indices[start_idx:start_idx + batch_size]
                
                batch_u_txt = [u_raw_texts[i] for i in batch_idx]
                batch_v_txt = [v_raw_texts[i] for i in batch_idx]
                
                batch_u_z = u_zones[batch_idx].to(self.device)
                batch_v_z = v_zones[batch_idx].to(self.device)
                batch_y = labels[batch_idx].to(self.device)

                optimizer_joint.zero_grad()

                u_inputs = self.encoder.tokenize(batch_u_txt)
                u_inputs = {k: v.to(self.device) for k, v in u_inputs.items()}
                u_emb = self.encoder(u_inputs)['sentence_embedding']

                v_inputs = self.encoder.tokenize(batch_v_txt)
                v_inputs = {k: v.to(self.device) for k, v in v_inputs.items()}
                v_emb = self.encoder(v_inputs)['sentence_embedding']

                logits = self.classifier(u_emb, batch_u_z, v_emb, batch_v_z)
                loss = criterion(logits, batch_y)
                loss.backward()
                optimizer_joint.step()
                total_loss += loss.item()

        self.encoder.eval()
        self.classifier.eval()

        # Calculate metrics
        with torch.no_grad():
            u_embs, v_embs = [], []
            for start_idx in range(0, num_samples, batch_size):
                end_idx = min(start_idx + batch_size, num_samples)
                batch_u_txt = u_raw_texts[start_idx:end_idx]
                batch_v_txt = v_raw_texts[start_idx:end_idx]
                
                u_inputs = self.encoder.tokenize(batch_u_txt)
                u_inputs = {k: v.to(self.device) for k, v in u_inputs.items()}
                u_embs.append(self.encoder(u_inputs)['sentence_embedding'])

                v_inputs = self.encoder.tokenize(batch_v_txt)
                v_inputs = {k: v.to(self.device) for k, v in v_inputs.items()}
                v_embs.append(self.encoder(v_inputs)['sentence_embedding'])

            u_tensors_eval = torch.cat(u_embs)
            v_tensors_eval = torch.cat(v_embs)
            
            logits = self.classifier(u_tensors_eval, u_zones.to(self.device), v_tensors_eval, v_zones.to(self.device))
            probs = torch.sigmoid(logits).cpu().numpy().squeeze()
            y_true = labels.cpu().numpy().squeeze()
            
        preds = (probs > 0.5).astype(int)

        try:
            fpr_vals, tpr_vals, roc_thresholds = roc_curve(y_true, probs)
            youden_idx = int(np.argmax(tpr_vals - fpr_vals))
            optimal_threshold = float(roc_thresholds[youden_idx])
            self.pairwise_threshold = float(np.clip(optimal_threshold, 0.01, 0.99))
        except Exception as exc:
            logger.warning("ROC threshold computation failed (%s); falling back to 0.5.", exc)
            self.pairwise_threshold = 0.5
        
        return {
            "pairwise_examples": len(dataset),
            "positive_examples": int(y_true.sum()),
            "negative_examples": int(len(y_true) - y_true.sum()),
            "validation_precision": float(precision_score(y_true, preds, zero_division=0)),
            "validation_recall": float(recall_score(y_true, preds, zero_division=0)),
            "validation_f1": float(f1_score(y_true, preds, zero_division=0)),
            "validation_average_precision": float(average_precision_score(y_true, probs)),
        }

    def _build_pairwise_clusters(self, ml_df: pd.DataFrame, row_vectors: torch.Tensor) -> np.ndarray:
        import torch
        import numpy as np
        import pandas as pd

        course_array = np.full(len(ml_df), -1, dtype=int)
        global_course_id = 0

        # Group data into isolated temporal buckets
        buckets = {}
        for pos in range(len(ml_df)):
            row = ml_df.iloc[pos]
            bucket_key = (row["dispatch_date"], row["dispatch_time_key"])
            buckets.setdefault(bucket_key, []).append(pos)

        self.classifier.eval()
        
        # Track categorical attributes and zone arrays
        zone_ints = [self.zone_to_idx.get(z, 0) for z in ml_df["Zone"]]
        zone_strings = ml_df["Zone"].tolist()
        zone_indices = torch.tensor(zone_ints, dtype=torch.long)

        # Dynamically discover the exact text column name for pickup addresses
        address_col = None
        for col in ["Depart", "adresse", "pickup_address", "Address", "address"]:
            if col in ml_df.columns:
                address_col = col
                break

        for bucket_key, nodes in buckets.items():
            if not nodes:
                continue
            
            if len(nodes) == 1:
                course_array[nodes[0]] = global_course_id
                global_course_id += 1
                continue

            taxis = []
            assigned = set()

            # ─── PASS 1: HUMAN HEURISTIC — IDENTICAL ADDRESS EXTRACTION ───
            address_groups = {}
            for n in nodes:
                # Group by raw address string; fallback to stringified embedding if column missing
                addr_val = ml_df.iloc[n][address_col] if address_col else tuple(row_vectors[n].numpy().round(4))
                address_groups.setdefault(addr_val, []).append(n)
            
            leftover_nodes = []
            
            for addr_val, group_nodes in address_groups.items():
                if len(group_nodes) >= 2:
                    # Cleanly pack identical pickup locations into complete vehicles of 4
                    while len(group_nodes) >= 4:
                        chunk = group_nodes[:4]
                        taxis.append(chunk)
                        for n in chunk:
                            assigned.add(n)
                        group_nodes = group_nodes[4:]
                    
                    # Track remainder passengers to blend with their respective zones in Pass 2
                    if len(group_nodes) > 0:
                        leftover_nodes.extend(group_nodes)
                else:
                    leftover_nodes.extend(group_nodes)

            # ─── PASS 2: HUMAN HEURISTIC — ZONE HOMOGENEITY SEGREGATION ───
            zone_groups = {}
            for n in leftover_nodes:
                if n not in assigned:
                    zone_groups.setdefault(zone_strings[n], []).append(n)
            
            for zone_name, zone_nodes in zone_groups.items():
                # If the entire zone cohort cleanly fits into a single taxi vehicle
                if len(zone_nodes) <= self.max_passengers:
                    taxis.append(zone_nodes)
                    for n in zone_nodes:
                        assigned.add(n)
                else:
                    # ─── PASS 3: MODEL OPTIMIZATION WITHIN ZONE BOUNDARIES ───
                    # Generate candidate combinations exclusively within this specific zone
                    zone_pairs = []
                    for i in range(len(zone_nodes)):
                        for j in range(i + 1, len(zone_nodes)):
                            zone_pairs.append((zone_nodes[i], zone_nodes[j]))
                    
                    if zone_pairs:
                        left_idx = [p[0] for p in zone_pairs]
                        right_idx = [p[1] for p in zone_pairs]
                        
                        u_texts = row_vectors[left_idx].to(self.device)
                        v_texts = row_vectors[right_idx].to(self.device)
                        u_zones = zone_indices[left_idx].to(self.device)
                        v_zones = zone_indices[right_idx].to(self.device)
                        
                        with torch.no_grad():
                            outputs = self.classifier(u_texts, u_zones, v_texts, v_zones)
                            probs = torch.sigmoid(outputs).cpu().numpy().flatten()
                        
                        # Sort pairs descending by embedding/route similarity
                        scored_pairs = sorted(zip(probs, zone_pairs), key=lambda x: x[0], reverse=True)
                        
                        local_assigned = set()
                        local_taxis = {}
                        next_local_id = 0
                        
                        # Apply priority pairing restricted entirely within this zone boundary
                        for prob, (u, v) in scored_pairs:
                            u_in = u in local_assigned
                            v_in = v in local_assigned
                            
                            if not u_in and not v_in:
                                local_taxis[next_local_id] = [u, v]
                                local_assigned.add(u)
                                local_assigned.add(v)
                                next_local_id += 1
                            elif u_in and not v_in:
                                for tid, members in local_taxis.items():
                                    if u in members and len(members) < self.max_passengers:
                                        members.append(v)
                                        local_assigned.add(v)
                                        break
                            elif not u_in and v_in:
                                for tid, members in local_taxis.items():
                                    if v in members and len(members) < self.max_passengers:
                                        members.append(u)
                                        local_assigned.add(u)
                                        break
                        
                        # Commit the optimized intra-zone taxis
                        for tid, members in local_taxis.items():
                            taxis.append(members)
                            for n in members:
                                assigned.add(n)
                                
                    # Collect and chunk any outstanding singletons left over in this zone
                    zone_leftovers = [n for n in zone_nodes if n not in assigned]
                    while zone_leftovers:
                        chunk = zone_leftovers[:self.max_passengers]
                        taxis.append(chunk)
                        for n in chunk:
                            assigned.add(n)
                        zone_leftovers = zone_leftovers[self.max_passengers:]

            # ─── PASS 4: FALLBACK SAFETY NET ───
            final_leftovers = [n for n in nodes if n not in assigned]
            while final_leftovers:
                chunk = final_leftovers[:self.max_passengers]
                taxis.append(chunk)
                final_leftovers = final_leftovers[self.max_passengers:]

            # ─── 5. COMMIT THE STRUCTURAL TAXIS TO GLOBAL COURSE IDS ───
            for members in taxis:
                if members:
                    for pos in members:
                        course_array[pos] = global_course_id
                    global_course_id += 1

        return course_array

    def _build_dispatch_outputs(self, ml_df: pd.DataFrame, clusters: np.ndarray) -> Tuple[pd.DataFrame, pd.DataFrame]:
        dispatch_df = pd.DataFrame({
            "Course_ID": clusters,
            "Passenger": ml_df["Nom - Prénom"].values,
            "Zone": ml_df["Zone"].values,
            "Pickup_Location": ml_df["Ramassage"].values,
            "Dropoff_Location": ml_df["Destination"].values,
            "Hour": ml_df["hour"].values,
            "Weekday": ml_df["weekday"].values,
            "Route": ml_df["Ramassage"].values + " → " + ml_df["Destination"].values,
        }).sort_values("Course_ID").reset_index(drop=True)

        course_stats = dispatch_df.groupby("Course_ID").agg({
            "Passenger": "count",
            "Zone": lambda x: ", ".join(sorted(set(x))),
            "Route": lambda x: x.value_counts().index[0],
            "Hour": ["min", "max", "mean"],
            "Pickup_Location": "nunique",
            "Dropoff_Location": "nunique",
        }).round(1)
        course_stats.columns = ["Passengers", "Zones", "Main_Route", "Hour_Min", "Hour_Max", "Hour_Avg", "Pickups", "Dropoffs"]
        return dispatch_df, course_stats

    def _export_to_styled_excel(self, df: pd.DataFrame, output_path: str):
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data_Export")
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        tab = Table(displayName="ExportedData", ref=ws.dimensions)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col[:50] if cell.value] + [0])
            for cell in col: cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = self.reference_row_height
        wb.save(output_path)

    def _export_dispatch_workbook(self, input_path: Path, cleaned_df: pd.DataFrame, clusters: np.ndarray, output_path: str):
        template_path = Path("data/Historique.xlsx")
        try:
            twb = openpyxl.load_workbook(template_path)
            t_ws = twb[twb.sheetnames[0]]
            t_data = self._extract_template_metadata(t_ws)
        except Exception:
            t_data = (None, None, None, None, None)

        t_header_values, t_header_styles, t_header_height, t_data_styles, t_data_height = t_data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agents par Taxi"

        header_values = list(t_header_values) if isinstance(t_header_values, list) and t_header_values else list(cleaned_df.columns)
        for col_idx, value in enumerate(header_values, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            if isinstance(t_header_styles, dict) and col_idx in t_header_styles: self._apply_cell_style(cell, t_header_styles[col_idx])

        if t_header_height is not None: ws.row_dimensions[1].height = t_header_height

        ml_df = self._prepare_ml_dataframe(cleaned_df)
        self._render_template_layout(ws, cleaned_df, ml_df, clusters, use_sheet_as_template=False)
        wb.save(output_path)

    def _render_template_layout(self, ws, cleaned_df: pd.DataFrame, ml_df: pd.DataFrame, clusters: np.ndarray, use_sheet_as_template: bool = True):
        header_values = list(cleaned_df.columns)
        data_styles = {}
        self._apply_reference_dimensions(ws, len(header_values))

        header_to_source = self._resolve_display_source_map(cleaned_df, header_values)
        for idx, src in enumerate(header_to_source, start=1):
            if src and self._normalize_colname(src) == "taxi": header_values[idx - 1] = "TAXI"

        taxi_col_idx = self._find_taxi_column(header_values)
        self._apply_export_header_style(ws, 1, header_values, taxi_col_idx)

        white_fill, orange_fill = self._detect_course_fills(ws, taxi_col_idx)
        blocks = self._build_course_blocks(cleaned_df, ml_df, clusters)
        source_columns = set(cleaned_df.columns)
        output_row = 2
        prev_time_key = None
        color_toggle = False

        for _, (course_id, _, course_time_key, row_indices) in enumerate(blocks):
            if prev_time_key is not None and course_time_key != prev_time_key:
                self._apply_export_header_style(ws, output_row, header_values, taxi_col_idx)
                ws.row_dimensions[output_row].height = self.reference_row_height
                output_row += 1

            block_fill = orange_fill if color_toggle else white_fill
            block_start = output_row

            for source_row in row_indices:
                for col_idx, header in enumerate(header_values, start=1):
                    src_col = header_to_source[col_idx - 1] if col_idx - 1 < len(header_to_source) else None
                    value = cleaned_df.at[source_row, src_col] if src_col in source_columns else None
                    if taxi_col_idx and col_idx == taxi_col_idx: value = None
                    
                    cell = ws.cell(row=output_row, column=col_idx, value=value)
                    if taxi_col_idx is None or col_idx != taxi_col_idx: cell.fill = copy(block_fill)
                    try: cell.font = Font(size=self.reference_font_size)
                    except Exception: pass
                ws.row_dimensions[output_row].height = self.reference_row_height
                output_row += 1

            if taxi_col_idx and row_indices:
                taxi_cell = ws.cell(row=block_start, column=taxi_col_idx, value=f"Taxi_{course_id + 1}")
                self._apply_font_color(taxi_cell, "FFFF0000", bold=False)
                taxi_cell.font = Font(color="FFFF0000", size=self.reference_font_size)
                if output_row - 1 > block_start:
                    ws.merge_cells(start_row=block_start, start_column=taxi_col_idx, end_row=output_row - 1, end_column=taxi_col_idx)

            prev_time_key = course_time_key
            color_toggle = not color_toggle

    def _extract_template_metadata(self, ws):
        max_col = ws.max_column
        header_values = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, max_col + 1)]
        h_styles, d_styles = {}, {}
        for col_idx in range(1, max_col + 1):
            try: h_styles[col_idx] = self._extract_cell_style(ws.cell(row=1, column=col_idx))
            except Exception: pass
            if ws.max_row >= 2:
                try: d_styles[col_idx] = self._extract_cell_style(ws.cell(row=2, column=col_idx))
                except Exception: pass
        return header_values, h_styles, ws.row_dimensions[1].height, d_styles, (ws.row_dimensions[2].height if ws.max_row >= 2 else None)

    def _extract_cell_style(self, cell):
        return {attr: getattr(cell, attr, None) for attr in ("font", "fill", "border", "alignment", "number_format") if getattr(cell, attr, None) is not None}

    def _apply_cell_style(self, cell, style_parts):
        for attr, val in style_parts.items():
            try: setattr(cell, attr, val)
            except Exception: pass

    def _apply_reference_dimensions(self, ws, num_columns: int):
        for col_idx in range(1, num_columns + 1):
            width = self.reference_col_widths[col_idx - 1] if col_idx - 1 < len(self.reference_col_widths) else 18.0
            try: ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            except Exception: pass

    def _resolve_display_source_map(self, cleaned_df: pd.DataFrame, header_values: List[object]) -> List[Optional[str]]:
        norm_src = {self._normalize_colname(c): c for c in cleaned_df.columns}
        aliases = {
            "date": ["date"], "numero": ["numero", "matricule"], "nomprenom": ["nomprenom", "name", "passenger"],
            "ramassage": ["ramassage", "operation", "pickup"], "destination": ["destination", "dropoff"],
            "heure": ["heure", "time", "arrivaltime"], "taxi": ["taxi"]
        }
        res = []
        for header in header_values:
            if header in cleaned_df.columns: res.append(header); continue
            h_key = self._normalize_colname(header)
            src = next((norm_src[self._normalize_colname(cand)] for cand in aliases.get(h_key, [h_key]) if self._normalize_colname(cand) in norm_src), norm_src.get(h_key))
            res.append(src)
        return res

    def _apply_font_color(self, cell, color: str, bold: Optional[bool] = None):
        try:
            font = copy(cell.font)
            font.color = color
            if bold is not None: font.bold = bold
            cell.font = font
        except Exception: pass

    def _apply_export_header_style(self, ws, row: int, header_values: List[str], taxi_col_idx: Optional[int]):
        for col_idx, header in enumerate(header_values, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value, cell.fill = header, PatternFill(fill_type="solid", fgColor=self.reference_header_fill)
            font_color = self.reference_taxi_header_text if taxi_col_idx == col_idx else self.reference_header_text
            cell.font, cell.alignment = Font(color=font_color, bold=True, size=self.reference_font_size), Alignment(horizontal="center", vertical="center")

    def _build_course_blocks(self, cleaned_df: pd.DataFrame, ml_df: pd.DataFrame, clusters: np.ndarray) -> List[Tuple]:
        cols = self._resolve_required_columns(cleaned_df)
        heure_col = cols["Heure"]
        blocks = []
        for course_id in sorted(np.unique(clusters).tolist()):
            row_indices = np.where(clusters == course_id)[0].tolist()
            if row_indices:
                idx = row_indices[0]
                h = int(ml_df.iloc[idx]["hour"])
                r_heure = cleaned_df.iloc[idx][heure_col] if heure_col in cleaned_df.columns else ""
                blocks.append((int(course_id), h, str(r_heure).strip(), row_indices))
        blocks.sort(key=lambda item: (item[1], item[2], item[0]))
        return blocks

    def _find_taxi_column(self, headers: List[object]) -> Optional[int]:
        norm = [self._normalize_colname(str(h)) for h in headers]
        for idx, key in enumerate(norm, start=1):
            if key == "taxi": return idx
        for idx, key in enumerate(norm, start=1):
            if "taxi" in key: return idx
        return None

    def _detect_course_fills(self, ws, taxi_col_idx: Optional[int]):
        return PatternFill(fill_type="solid", fgColor=self.reference_course_white), PatternFill(fill_type="solid", fgColor=self.reference_course_orange)

    def export_training_artifacts(self, input_path: str, output_dir: str = "data") -> Dict[str, object]:
        source_path, output_path = Path(input_path), Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        raw_df = self._load_structured_input(source_path)
        cleaned_df = self._remove_repeated_headers(raw_df)
        row_to_group = MergeGroupReader(source_path).read()
        ml_df = self._prepare_ml_dataframe(cleaned_df, row_to_group)
        
        row_vectors = self._vectorize_addresses(ml_df, fit=True)
        pairwise_stats = self._fit_pairwise_model(ml_df, row_vectors, row_to_group)
        
        # Recompute vectors immediately after training so they leverage the fine-tuned encoder
        logger.info("Recomputing embeddings with the newly fine-tuned encoder...")
        row_vectors = self._vectorize_addresses(ml_df, fit=False)
        clusters = self._build_pairwise_clusters(ml_df, row_vectors)
        dispatch_df, course_stats = self._build_dispatch_outputs(ml_df, clusters)

        export_cleaned_df = cleaned_df.drop(columns=[col for col in ["matricule", "_excel_row", "_sheet_name"] if col in cleaned_df.columns], errors='ignore')

        export_cleaned_df.to_csv(output_path / "taxi_data_cleaned.csv", index=False)
        dispatch_df.to_csv(output_path / "final_course_dispatch_geographic.csv", index=False)
        course_stats.to_csv(output_path / "course_summary_geographic.csv")
        self._save_artifacts(str(output_path))

        report = {
            "source_file": str(source_path),
            "rows_raw": int(len(raw_df)),
            "rows_cleaned": int(len(cleaned_df)),
            "rows_training": int(len(ml_df)),
            "courses_created": int(dispatch_df["Course_ID"].nunique()),
            "pairwise_threshold": float(self.pairwise_threshold),
            **pairwise_stats
        }
        
        report_json = output_path / "model_training_report.json"
        report_json.with_name(report_json.name + ".tmp").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        os.replace(report_json.with_name(report_json.name + ".tmp"), report_json)
        return report

if __name__ == "__main__":
    converter = FileConverter(trip_type="ramassage")
    try:
        print("Converter initialized and ready.")
    except Exception as e:
        logger.error(f"Error: {e}")